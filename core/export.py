"""Export — Rapport Excel professionnel avec graphiques natifs.

Couleurs alignées sur le dashboard Streamlit :
  PTF = bleu #2E75B6, Benchmark = orange #FF6B35, négatif = rouge #E74C3C
"""

import io
from datetime import date

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

from core.portfolio import Portfolio
from core import metrics
from core.contribution import performance_contribution, risk_contribution

# ── Palette (identique au Streamlit) ────────────────────
_PTF = "2E75B6"
_BENCH = "FF6B35"
_NAVY = "1B3A5C"
_LIGHT_BLUE = "D6E4F0"
_GREEN = "27AE60"
_LIGHT_GREEN = "D5F5E3"
_RED = "E74C3C"
_LIGHT_RED = "FADBD8"
_PURPLE = "9B59B6"
_TEAL = "1ABC9C"
_GREY_BG = "F7F8FA"
_GREY_BORDER = "D5D8DC"
_DARK = "1B2A3D"
_WHITE = "FFFFFF"

# ── Styles réutilisables ────────────────────────────────
_F_TITLE = Font(bold=True, color=_NAVY, size=16, name="Calibri")
_F_SECTION = Font(bold=True, color=_WHITE, size=11, name="Calibri")
_F_SUBSECTION = Font(bold=True, color=_NAVY, size=11, name="Calibri")
_F_BODY = Font(color=_DARK, size=10, name="Calibri")
_F_METRIC = Font(bold=True, color=_NAVY, size=10, name="Calibri")
_F_HEADER = Font(bold=True, color=_WHITE, size=10, name="Calibri")

_FILL_HEADER = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type="solid")
_FILL_SUBHEADER = PatternFill(start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid")
_FILL_ZEBRA = PatternFill(start_color=_GREY_BG, end_color=_GREY_BG, fill_type="solid")
_FILL_GOOD = PatternFill(start_color=_LIGHT_GREEN, end_color=_LIGHT_GREEN, fill_type="solid")
_FILL_BAD = PatternFill(start_color=_LIGHT_RED, end_color=_LIGHT_RED, fill_type="solid")

_BORDER_THIN = Border(bottom=Side(style="thin", color=_GREY_BORDER))
_BORDER_HEADER = Border(bottom=Side(style="medium", color=_NAVY))

# Chart dimensions (consistent across all sheets)
_CHART_W = 32   # ~16 columns
_CHART_H = 16   # ~30 rows in Excel
_CHART_GAP = 33  # rows between stacked charts


# ── Helpers ─────────────────────────────────────────────

def _header_row(ws, row, cols: list[str]):
    for i, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=label)
        c.font = _F_HEADER
        c.fill = _FILL_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER_HEADER


def _zebra(ws, r_start, r_end, ncols):
    for r in range(r_start, r_end + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            if not cell.font or cell.font == Font():
                cell.font = _F_BODY
            cell.border = _BORDER_THIN
            if (r - r_start) % 2 == 1:
                if not cell.fill or cell.fill == PatternFill():
                    cell.fill = _FILL_ZEBRA


def _title(ws, row, text, ncols):
    c = ws.cell(row=row, column=1, value=text)
    c.font = _F_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def _section(ws, row, text, ncols):
    """Blue background section header."""
    for i in range(1, ncols + 1):
        c = ws.cell(row=row, column=i)
        c.fill = _FILL_SUBHEADER
    ws.cell(row=row, column=1, value=text).font = _F_SUBSECTION
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)


def _auto_w(ws, ncols, min_w=14):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        mx = min_w
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value is not None:
                    mx = max(mx, len(str(cell.value)) + 3)
        ws.column_dimensions[letter].width = min(mx, 42)


def _metric_row(ws, row, label, ptf_v, bench_v, fmt="pct"):
    ws.cell(row=row, column=1, value=label).font = _F_BODY
    c2 = ws.cell(row=row, column=2)
    if ptf_v is not None:
        c2.value = ptf_v
        c2.number_format = {"pct": "0.00%", "num": "0.00", "int": "0"}[fmt]
        c2.font = _F_METRIC
    else:
        c2.value = "—"
        c2.font = _F_BODY
    c3 = ws.cell(row=row, column=3)
    if bench_v is not None:
        c3.value = bench_v
        c3.number_format = {"pct": "0.00%", "num": "0.00", "int": "0"}[fmt]
    else:
        c3.value = "—"
    c3.font = _F_BODY


def _style_line_chart(chart, title_text):
    """Apply consistent styling to a line chart."""
    chart.title = title_text
    chart.width = _CHART_W
    chart.height = _CHART_H
    chart.style = 10
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.numFmt = "MM/YYYY"


def _color_series(chart, idx, color, width=20000, dash=None):
    s = chart.series[idx]
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = width
    s.smooth = True
    if dash:
        s.graphicalProperties.line.dashStyle = dash


# ── Synthèse ────────────────────────────────────────────

def _build_synthese(wb, ptf, bench, rf, garch_ptf, kupiec_95, kupiec_99):
    ws = wb.active
    ws.title = "Synthèse"
    ws.sheet_properties.tabColor = _NAVY

    _title(ws, 1, "RAPPORT D'ANALYSE DE PORTEFEUILLE", 3)
    ws.cell(row=2, column=1,
            value=f"Généré le {date.today().strftime('%d/%m/%Y')}").font = _F_BODY

    # ── Performance
    row = 4
    _section(ws, row, "PERFORMANCE", 3)
    row += 1
    _header_row(ws, row, ["Métrique", "Portefeuille", "Benchmark"])
    data = [
        ("Rendement cumulé", metrics.cumulative_return(ptf),
         metrics.cumulative_return(bench), "pct"),
        ("CAGR", metrics.annualized_return(ptf),
         metrics.annualized_return(bench), "pct"),
        ("Alpha géométrique", metrics.geometric_alpha(ptf, bench), None, "pct"),
    ]
    for i, (lbl, pv, bv, fmt) in enumerate(data):
        _metric_row(ws, row + 1 + i, lbl, pv, bv, fmt)
    _zebra(ws, row + 1, row + len(data), 3)

    # ── Risque
    row = row + len(data) + 2
    _section(ws, row, "RISQUE", 3)
    row += 1
    _header_row(ws, row, ["Métrique", "Portefeuille", "Benchmark"])
    data = [
        ("Volatilité annualisée", metrics.annualized_volatility(ptf),
         metrics.annualized_volatility(bench), "pct"),
        ("Vol GARCH long terme",
         garch_ptf["long_run_vol"] if not pd.isna(garch_ptf["long_run_vol"]) else None,
         None, "pct"),
        ("Maximum Drawdown", metrics.max_drawdown(ptf),
         metrics.max_drawdown(bench), "pct"),
        ("Durée drawdown max (j)", metrics.max_drawdown_duration(ptf),
         metrics.max_drawdown_duration(bench), "int"),
        ("VaR 95% (1 jour)", metrics.var_95(ptf), metrics.var_95(bench), "pct"),
        ("CVaR 95%", metrics.cvar_95(ptf), metrics.cvar_95(bench), "pct"),
        ("VaR 99% (1 jour)", metrics.var_99(ptf), metrics.var_99(bench), "pct"),
        ("CVaR 99%", metrics.cvar_99(ptf), metrics.cvar_99(bench), "pct"),
        ("Tracking Error", metrics.tracking_error(ptf, bench), None, "pct"),
        ("Bêta", metrics.beta(ptf, bench), 1.0, "num"),
    ]
    for i, (lbl, pv, bv, fmt) in enumerate(data):
        _metric_row(ws, row + 1 + i, lbl, pv, bv, fmt)
    _zebra(ws, row + 1, row + len(data), 3)

    # ── Ratios ajustés
    row = row + len(data) + 2
    _section(ws, row, "RATIOS AJUSTÉS DU RISQUE", 3)
    row += 1
    _header_row(ws, row, ["Ratio", "Portefeuille", "Benchmark"])
    data = [
        ("Sharpe", metrics.sharpe_ratio(ptf, rf), metrics.sharpe_ratio(bench, rf), "num"),
        ("Sortino", metrics.sortino_ratio(ptf, rf), metrics.sortino_ratio(bench, rf), "num"),
        ("Information Ratio", metrics.information_ratio(ptf, bench), None, "num"),
        ("Calmar", metrics.calmar_ratio(ptf), metrics.calmar_ratio(bench), "num"),
    ]
    for i, (lbl, pv, bv, fmt) in enumerate(data):
        _metric_row(ws, row + 1 + i, lbl, pv, bv, fmt)
    _zebra(ws, row + 1, row + len(data), 3)

    # ── Kupiec
    row = row + len(data) + 2
    _section(ws, row, "BACKTESTING VaR — TEST DE KUPIEC", 3)
    row += 1
    _header_row(ws, row, ["Niveau", "p-value", "Verdict"])
    for i, (lbl, kup) in enumerate([("VaR 95%", kupiec_95), ("VaR 99%", kupiec_99)]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_BODY
        ws.cell(row=r, column=2, value=kup["p_value"]).font = _F_METRIC
        ws.cell(row=r, column=2).number_format = "0.0000"
        vc = ws.cell(row=r, column=3)
        if kup["model_adequate"]:
            vc.value = "Adéquat"
            vc.fill = _FILL_GOOD
            vc.font = Font(bold=True, color=_GREEN, size=10, name="Calibri")
        else:
            vc.value = "REJETÉ"
            vc.fill = _FILL_BAD
            vc.font = Font(bold=True, color=_RED, size=10, name="Calibri")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18


# ── Séries temporelles ──────────────────────────────────

def _build_series(wb, portfolio, ptf, bench):
    ws = wb.create_sheet("Séries temporelles")
    ws.sheet_properties.tabColor = _PTF

    _title(ws, 1, "SÉRIES TEMPORELLES", 5)
    row = 3
    _header_row(ws, row, ["Date", "PTF (base 100)", "Bench (base 100)",
                           "Drawdown PTF", "Drawdown Bench"])

    ptf_norm = ptf / ptf.iloc[0] * 100
    bench_norm = bench / bench.iloc[0] * 100
    dd_ptf = metrics.drawdown_series(ptf)
    dd_bench = metrics.drawdown_series(bench)

    idx = ptf_norm.index
    for i, dt in enumerate(idx):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=dt.date() if hasattr(dt, 'date') else dt)
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2, value=round(float(ptf_norm.iloc[i]), 2))
        ws.cell(row=r, column=3, value=round(float(bench_norm.iloc[i]), 2))
        ws.cell(row=r, column=4, value=round(float(dd_ptf.iloc[i]), 6))
        ws.cell(row=r, column=4).number_format = "0.00%"
        ws.cell(row=r, column=5, value=round(float(dd_bench.iloc[i]), 6))
        ws.cell(row=r, column=5).number_format = "0.00%"

    n = row + len(idx)
    _auto_w(ws, 5)
    dates_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=n)
    skip = max(1, len(idx) // 12)

    # Chart 1 — Performance cumulée
    c1 = LineChart()
    _style_line_chart(c1, "Performance cumulée (base 100)")
    c1.y_axis.title = "Valeur"
    for col in [2, 3]:
        c1.add_data(Reference(ws, min_col=col, min_row=row, max_row=n), titles_from_data=True)
    c1.set_categories(dates_ref)
    _color_series(c1, 0, _PTF, 22000)
    _color_series(c1, 1, _BENCH, 18000, "dash")
    c1.x_axis.tickLblSkip = skip
    ws.add_chart(c1, "G3")

    # Chart 2 — Drawdown (placé 33 lignes plus bas)
    c2 = LineChart()
    _style_line_chart(c2, "Drawdown")
    c2.y_axis.title = "Drawdown"
    c2.y_axis.numFmt = "0.0%"
    for col in [4, 5]:
        c2.add_data(Reference(ws, min_col=col, min_row=row, max_row=n), titles_from_data=True)
    c2.set_categories(dates_ref)
    _color_series(c2, 0, _RED, 16000)
    _color_series(c2, 1, "95A5A6", 12000, "dash")
    c2.x_axis.tickLblSkip = skip
    ws.add_chart(c2, f"G{3 + _CHART_GAP}")


# ── GARCH & GJR-GARCH ──────────────────────────────────

def _build_garch(wb, ptf, bench, garch_ptf, garch_bench):
    ws = wb.create_sheet("GARCH & GJR")
    ws.sheet_properties.tabColor = _PURPLE

    _title(ws, 1, "MODÈLES DE VOLATILITÉ CONDITIONNELLE", 4)

    # ── GARCH(1,1) params
    row = 3
    _section(ws, row, "GARCH(1,1) — Bollerslev (1986)", 4)
    row += 1
    _header_row(ws, row, ["Paramètre", "Portefeuille", "Benchmark", ""])
    params = [
        ("ω (omega)", garch_ptf["omega"], garch_bench["omega"], "0.000000"),
        ("α (réaction)", garch_ptf["alpha"], garch_bench["alpha"], "0.0000"),
        ("β (persistance)", garch_ptf["beta"], garch_bench["beta"], "0.0000"),
        ("Persistance (α+β)", garch_ptf["persistence"], garch_bench["persistence"], "0.0000"),
        ("Vol long terme (ann.)",
         garch_ptf["long_run_vol"] if not pd.isna(garch_ptf["long_run_vol"]) else None,
         garch_bench["long_run_vol"] if not pd.isna(garch_bench["long_run_vol"]) else None,
         "0.00%"),
        ("Vol daily J+1", garch_ptf["forecast_vol_1d"],
         garch_bench["forecast_vol_1d"], "0.000%"),
        ("Vol daily J+5 (moy.)", garch_ptf["forecast_vol_5d"],
         garch_bench["forecast_vol_5d"], "0.000%"),
        ("Vol daily J+10 (moy.)", garch_ptf["forecast_vol_10d"],
         garch_bench["forecast_vol_10d"], "0.000%"),
    ]
    for i, (lbl, pv, bv, nf) in enumerate(params):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_BODY
        for ci, val in [(2, pv), (3, bv)]:
            c = ws.cell(row=r, column=ci)
            if val is not None:
                c.value = float(val)
                c.number_format = nf
                c.font = _F_METRIC
            else:
                c.value = "—"
                c.font = _F_BODY
    _zebra(ws, row + 1, row + len(params), 3)

    # ── GJR-GARCH params
    row = row + len(params) + 2
    _section(ws, row, "GJR-GARCH(1,1) — Glosten-Jagannathan-Runkle (1993)", 4)
    row += 1
    _header_row(ws, row, ["Paramètre", "Portefeuille", "Benchmark", ""])

    gjr_ptf = metrics.fit_gjr_garch(ptf)
    gjr_bench = metrics.fit_gjr_garch(bench)

    gjr_params = [
        ("α (réaction)", gjr_ptf["alpha"], gjr_bench["alpha"], "0.0000"),
        ("γ (effet de levier)", gjr_ptf["gamma"], gjr_bench["gamma"], "0.0000"),
        ("β (persistance)", gjr_ptf["beta"], gjr_bench["beta"], "0.0000"),
        ("Persistance (α+β+γ/2)", gjr_ptf["persistence"], gjr_bench["persistence"], "0.0000"),
        ("Vol long terme (ann.)",
         gjr_ptf["long_run_vol"] if not pd.isna(gjr_ptf["long_run_vol"]) else None,
         gjr_bench["long_run_vol"] if not pd.isna(gjr_bench["long_run_vol"]) else None,
         "0.00%"),
        ("Vol daily J+1", gjr_ptf["forecast_vol_1d"],
         gjr_bench["forecast_vol_1d"], "0.000%"),
        ("Vol daily J+5 (moy.)", gjr_ptf["forecast_vol_5d"],
         gjr_bench["forecast_vol_5d"], "0.000%"),
        ("Vol daily J+10 (moy.)", gjr_ptf["forecast_vol_10d"],
         gjr_bench["forecast_vol_10d"], "0.000%"),
    ]
    for i, (lbl, pv, bv, nf) in enumerate(gjr_params):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_BODY
        for ci, val in [(2, pv), (3, bv)]:
            c = ws.cell(row=r, column=ci)
            if val is not None:
                c.value = float(val)
                c.number_format = nf
                c.font = _F_METRIC
            else:
                c.value = "—"
                c.font = _F_BODY
    _zebra(ws, row + 1, row + len(gjr_params), 3)

    # Leverage effect verdict
    r_verdict = row + len(gjr_params) + 1
    vc = ws.cell(row=r_verdict, column=1)
    if gjr_ptf["leverage_effect"]:
        vc.value = f"Effet de levier confirmé (γ = {gjr_ptf['gamma']:.4f} > 0)"
        vc.font = Font(bold=True, color=_GREEN, size=10, name="Calibri")
    else:
        vc.value = f"Pas d'effet de levier (γ = {gjr_ptf['gamma']:.4f} ≤ 0)"
        vc.font = Font(italic=True, color="95A5A6", size=10, name="Calibri")
    ws.merge_cells(start_row=r_verdict, start_column=1,
                   end_row=r_verdict, end_column=3)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # ── Vol time series data
    vol_ptf = garch_ptf["conditional_vol"]
    vol_bench = garch_bench["conditional_vol"]
    gjr_vol_ptf = gjr_ptf["conditional_vol"]
    common = vol_ptf.index.intersection(vol_bench.index).intersection(gjr_vol_ptf.index)

    ds = r_verdict + 2
    _header_row(ws, ds, ["Date", "Vol daily GARCH PTF", "Vol daily GARCH Bench",
                          "Vol daily GJR PTF"])
    for i, dt in enumerate(common):
        r = ds + 1 + i
        ws.cell(row=r, column=1, value=dt.date() if hasattr(dt, 'date') else dt)
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2, value=round(float(vol_ptf.loc[dt]), 6))
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=3, value=round(float(vol_bench.loc[dt]), 6))
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=4, value=round(float(gjr_vol_ptf.loc[dt]), 6))
        ws.cell(row=r, column=4).number_format = "0.00%"

    n_data = ds + len(common)
    skip = max(1, len(common) // 12)

    # Chart GARCH
    cg = LineChart()
    _style_line_chart(cg, "Volatilité conditionnelle GARCH(1,1) — daily")
    cg.y_axis.title = "Vol daily"
    cg.y_axis.numFmt = "0.00%"
    dates_ref = Reference(ws, min_col=1, min_row=ds + 1, max_row=n_data)
    for col in [2, 3]:
        cg.add_data(Reference(ws, min_col=col, min_row=ds, max_row=n_data),
                    titles_from_data=True)
    cg.set_categories(dates_ref)
    _color_series(cg, 0, _PTF, 20000)
    _color_series(cg, 1, _BENCH, 16000, "dash")
    cg.x_axis.tickLblSkip = skip
    ws.add_chart(cg, "F2")

    # Chart GJR-GARCH (33 rows below)
    cj = LineChart()
    _style_line_chart(cj, "Volatilité conditionnelle GJR-GARCH(1,1) — daily")
    cj.y_axis.title = "Vol daily"
    cj.y_axis.numFmt = "0.00%"
    for col in [2, 4]:
        cj.add_data(Reference(ws, min_col=col, min_row=ds, max_row=n_data),
                    titles_from_data=True)
    cj.set_categories(dates_ref)
    _color_series(cj, 0, _PTF, 18000)
    _color_series(cj, 1, _PURPLE, 20000)
    cj.x_axis.tickLblSkip = skip
    ws.add_chart(cj, f"F{2 + _CHART_GAP}")


# ── Rolling Sharpe ──────────────────────────────────────

def _build_rolling_sharpe(wb, ptf, bench, rf):
    ws = wb.create_sheet("Rolling Sharpe")
    ws.sheet_properties.tabColor = _PTF

    _title(ws, 1, "ROLLING SHARPE RATIO (1 AN)", 3)
    row = 3
    _header_row(ws, row, ["Date", "Sharpe PTF", "Sharpe Benchmark"])

    rs_ptf = metrics.rolling_sharpe(ptf, risk_free_rate=rf)
    rs_bench = metrics.rolling_sharpe(bench, risk_free_rate=rf)
    common = rs_ptf.index.intersection(rs_bench.index)

    for i, dt in enumerate(common):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=dt.date() if hasattr(dt, 'date') else dt)
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2, value=round(float(rs_ptf.loc[dt]), 4))
        ws.cell(row=r, column=2).number_format = "0.00"
        ws.cell(row=r, column=3, value=round(float(rs_bench.loc[dt]), 4))
        ws.cell(row=r, column=3).number_format = "0.00"

    n = row + len(common)
    _auto_w(ws, 3)

    ch = LineChart()
    _style_line_chart(ch, "Rolling Sharpe Ratio (1 an)")
    ch.y_axis.title = "Sharpe"
    dates_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=n)
    for col in [2, 3]:
        ch.add_data(Reference(ws, min_col=col, min_row=row, max_row=n),
                    titles_from_data=True)
    ch.set_categories(dates_ref)
    _color_series(ch, 0, _PTF, 20000)
    _color_series(ch, 1, _BENCH, 16000, "dash")
    ch.x_axis.tickLblSkip = max(1, len(common) // 12)
    ws.add_chart(ch, "E3")


# ── Corrélations ────────────────────────────────────────

def _build_correlations(wb, portfolio):
    ws = wb.create_sheet("Corrélations")
    ws.sheet_properties.tabColor = _GREEN

    corr = portfolio.correlation_matrix
    nc = len(corr.columns)

    _title(ws, 1, "MATRICE DE CORRÉLATION", nc + 1)

    row = 3
    ws.cell(row=row, column=1, value="")
    for i, t in enumerate(corr.columns, 2):
        ws.cell(row=row, column=i, value=t)
    _header_row(ws, row, [""] + list(corr.columns))

    for i, ticker in enumerate(corr.index):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=ticker).font = _F_SUBSECTION
        for j, ct in enumerate(corr.columns):
            c_idx = j + 2
            val = float(corr.loc[ticker, ct])
            cell = ws.cell(row=r, column=c_idx, value=round(val, 2))
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            av = abs(val)
            if ticker == ct:
                cell.fill = _FILL_SUBHEADER
                cell.font = _F_METRIC
            elif av > 0.8:
                cell.fill = _FILL_BAD
                cell.font = Font(bold=True, color=_RED, size=10, name="Calibri")
            elif av < 0.3:
                cell.fill = _FILL_GOOD
                cell.font = Font(color=_GREEN, size=10, name="Calibri")
            else:
                cell.font = _F_BODY

    _auto_w(ws, nc + 1)


# ── Contribution ────────────────────────────────────────

def _build_contribution(wb, portfolio):
    ws = wb.create_sheet("Contribution")
    ws.sheet_properties.tabColor = _BENCH

    _title(ws, 1, "ANALYSE DE CONTRIBUTION", 5)

    perf_c = performance_contribution(portfolio)
    risk_c = risk_contribution(portfolio)
    risk_abs = risk_c.abs()

    row = 3
    _header_row(ws, row, ["Actif", "Poids", "Contrib. perf.", "Contrib. risque", "Risque (abs)"])

    tickers = list(perf_c.index)
    for i, t in enumerate(tickers):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=t).font = _F_SUBSECTION
        for ci, val in [(2, portfolio.weights[t]), (3, float(perf_c[t])),
                         (4, float(risk_c[t])), (5, float(risk_abs[t]))]:
            cell = ws.cell(row=r, column=ci, value=round(val, 6))
            cell.number_format = "0.00%"
            cell.font = _F_METRIC
        # Color perf contribution
        pc = ws.cell(row=r, column=3)
        pc.fill = _FILL_GOOD if float(perf_c[t]) >= 0 else _FILL_BAD

    n = row + len(tickers)
    _auto_w(ws, 5)

    # Bar chart — contribution perf (col H, row 3)
    bar = BarChart()
    bar.type = "bar"
    bar.title = "Contribution à la performance"
    bar.x_axis.numFmt = "0.0%"
    bar.width = 24
    bar.height = 14
    bar.style = 10
    bar.add_data(Reference(ws, min_col=3, min_row=row, max_row=n), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=row + 1, max_row=n))

    s = bar.series[0]
    s.graphicalProperties.solidFill = _RED  # default = red, overridden per bar
    for i, t in enumerate(tickers):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = _PTF if float(perf_c[t]) >= 0 else _RED
        s.data_points.append(pt)
    bar.legend = None
    ws.add_chart(bar, "H3")

    # Pie chart — risque (réduit + décalé pour ne pas chevaucher le titre)
    pie = PieChart()
    pie.title = "Décomposition du risque"
    pie.width = 18
    pie.height = 12
    pie.style = 10
    pie.add_data(Reference(ws, min_col=5, min_row=row, max_row=n), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=row + 1, max_row=n))
    pie.series[0].data_labels = DataLabelList()
    pie.series[0].data_labels.showPercent = True
    pie.series[0].data_labels.showCatName = True

    # Monochrome blue gradient like Streamlit
    max_rc = risk_abs.max() if risk_abs.max() > 0 else 1
    for i, t in enumerate(tickers):
        pt = DataPoint(idx=i)
        ratio = float(risk_abs[t]) / max_rc
        # Interpolate from light blue to navy
        r_val = int(0xD6 + (0x1B - 0xD6) * ratio)
        g_val = int(0xE4 + (0x3A - 0xE4) * ratio)
        b_val = int(0xF0 + (0x5C - 0xF0) * ratio)
        pt.graphicalProperties.solidFill = f"{r_val:02X}{g_val:02X}{b_val:02X}"
        pie.series[0].data_points.append(pt)
    ws.add_chart(pie, "H37")


# ── Distribution ────────────────────────────────────────

def _build_distribution(wb, ptf, bench):
    ws = wb.create_sheet("Distribution")
    ws.sheet_properties.tabColor = _RED

    _title(ws, 1, "DISTRIBUTION DES RENDEMENTS & VALUE AT RISK", 6)

    rets = metrics.daily_returns(ptf)
    var95 = metrics.var_95(ptf)
    cvar95 = metrics.cvar_95(ptf)
    var99 = metrics.var_99(ptf)
    cvar99 = metrics.cvar_99(ptf)

    # Summary
    row = 3
    _section(ws, row, "STATISTIQUES VaR (HORIZON 1 JOUR)", 2)
    row += 1
    _header_row(ws, row, ["Statistique", "Valeur"])
    for i, (lbl, val) in enumerate([
        ("VaR 95%", var95), ("CVaR 95%", cvar95),
        ("VaR 99%", var99), ("CVaR 99%", cvar99),
    ]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_BODY
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = "0.00%"
        c.font = _F_METRIC
        c.fill = _FILL_BAD

    # Histogram data
    bins = np.linspace(float(rets.min()), float(rets.max()), 51)
    counts, edges = np.histogram(rets.values, bins=bins)
    centers = (edges[:-1] + edges[1:]) / 2

    hs = row + 7
    _header_row(ws, hs, ["Bin", "Fréquence"])
    for i in range(len(counts)):
        r = hs + 1 + i
        ws.cell(row=r, column=1, value=round(float(centers[i]), 6))
        ws.cell(row=r, column=1).number_format = "0.00%"
        ws.cell(row=r, column=2, value=int(counts[i]))

    n_hist = hs + len(counts)
    _auto_w(ws, 2)

    # Histogram chart (placed to the right of data, row 3)
    ch = BarChart()
    ch.type = "col"
    ch.title = "Distribution des rendements quotidiens"
    ch.x_axis.title = "Rendement"
    ch.y_axis.title = "Fréquence"
    ch.width = _CHART_W
    ch.height = _CHART_H
    ch.style = 10
    ch.gapWidth = 0
    ch.add_data(Reference(ws, min_col=2, min_row=hs, max_row=n_hist), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=hs + 1, max_row=n_hist))

    s = ch.series[0]
    s.graphicalProperties.solidFill = _PTF
    for i in range(len(counts)):
        if centers[i] <= var99:
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = _PURPLE
            s.data_points.append(pt)
        elif centers[i] <= var95:
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = _RED
            s.data_points.append(pt)

    ch.legend = None
    ch.x_axis.numFmt = "0.0%"
    ch.x_axis.tickLblSkip = max(1, len(counts) // 10)
    ws.add_chart(ch, "D3")


# ── API publique ────────────────────────────────────────

def _build_regimes(wb, ptf, ms_ptf):
    """Onglet Régimes de marché — Markov Switching."""
    ws = wb.create_sheet("Régimes")
    ws.sheet_properties.tabColor = _PURPLE

    _title(ws, 1, "DÉTECTION DE RÉGIME — MARKOV SWITCHING", 4)

    # Regime characteristics
    row = 3
    _section(ws, row, "CARACTÉRISTIQUES DES RÉGIMES", 4)
    row += 1
    _header_row(ws, row, ["Régime", "Rendement ann.", "Volatilité ann.", "Durée moy. (j)"])

    calm = ms_ptf["regimes"]["calme"]
    stress = ms_ptf["regimes"]["stress"]
    for i, (lbl, reg, dur) in enumerate([
        ("Calme", calm, ms_ptf["expected_duration_calm"]),
        ("Stress", stress, ms_ptf["expected_duration_stress"]),
    ]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_SUBSECTION
        ws.cell(row=r, column=2, value=reg["mean_ann"])
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=2).font = _F_METRIC
        ws.cell(row=r, column=3, value=reg["vol_ann"])
        ws.cell(row=r, column=3).number_format = "0.00%"
        ws.cell(row=r, column=3).font = _F_METRIC
        ws.cell(row=r, column=4, value=round(dur))
        ws.cell(row=r, column=4).font = _F_METRIC
        if lbl == "Stress":
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = _FILL_BAD

    # Current status
    row = row + 4
    _section(ws, row, "ÉTAT ACTUEL", 4)
    row += 1
    ws.cell(row=row, column=1, value="P(stress) actuelle").font = _F_BODY
    pc = ws.cell(row=row, column=2, value=ms_ptf["current_regime_proba"])
    pc.number_format = "0.0%"
    pc.font = _F_METRIC
    if ms_ptf["current_regime_proba"] >= 0.7:
        pc.fill = _FILL_BAD
    elif ms_ptf["current_regime_proba"] <= 0.3:
        pc.fill = _FILL_GOOD

    # Transition matrix
    row += 2
    _section(ws, row, "MATRICE DE TRANSITION", 4)
    row += 1
    _header_row(ws, row, ["", "→ Calme", "→ Stress", ""])
    trans = ms_ptf["transition_matrix"]
    si = ms_ptf["stress_regime"]
    ci = 1 - si
    for i, (lbl, from_idx) in enumerate([("Calme", ci), ("Stress", si)]):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=lbl).font = _F_SUBSECTION
        ws.cell(row=r, column=2, value=trans[from_idx][ci])
        ws.cell(row=r, column=2).number_format = "0.0000"
        ws.cell(row=r, column=2).font = _F_METRIC
        ws.cell(row=r, column=3, value=trans[from_idx][si])
        ws.cell(row=r, column=3).number_format = "0.0000"
        ws.cell(row=r, column=3).font = _F_METRIC

    # Time series data
    row = row + 4
    filtered = ms_ptf["filtered_proba_stress"]
    smoothed = ms_ptf["smoothed_proba_stress"]
    _header_row(ws, row, ["Date", "P(stress) filtrée", "P(stress) lissée", ""])

    for i, dt in enumerate(filtered.index):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=dt.date() if hasattr(dt, 'date') else dt)
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2, value=round(float(filtered.iloc[i]), 4))
        ws.cell(row=r, column=2).number_format = "0.0%"
        ws.cell(row=r, column=3, value=round(float(smoothed.iloc[i]), 4))
        ws.cell(row=r, column=3).number_format = "0.0%"

    n_data = row + len(filtered)
    _auto_w(ws, 4)

    # Chart — filtered probability
    ch = LineChart()
    _style_line_chart(ch, "Probabilité du régime stress")
    ch.y_axis.title = "P(stress)"
    ch.y_axis.numFmt = "0%"
    ch.y_axis.scaling.min = 0
    ch.y_axis.scaling.max = 1

    dates_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=n_data)
    for col in [2, 3]:
        ch.add_data(Reference(ws, min_col=col, min_row=row, max_row=n_data),
                    titles_from_data=True)
    ch.set_categories(dates_ref)
    _color_series(ch, 0, _RED, 16000)
    _color_series(ch, 1, _PURPLE, 14000, "dash")
    ch.x_axis.tickLblSkip = max(1, len(filtered) // 12)
    ws.add_chart(ch, "F3")


def generate_excel(portfolio: Portfolio, bench_prices: pd.Series,
                   risk_free_rate: float,
                   garch_ptf=None, garch_bench=None,
                   kupiec_95=None, kupiec_99=None,
                   ms_ptf=None) -> bytes:
    """Génère un rapport Excel professionnel multi-onglets."""
    wb = Workbook()
    ptf = portfolio.cumulative_prices

    common = ptf.index.intersection(bench_prices.index)
    ptf = ptf.loc[common]
    bench = bench_prices.loc[common]

    if garch_ptf is None:
        garch_ptf = metrics.fit_garch(ptf)
    if garch_bench is None:
        garch_bench = metrics.fit_garch(bench)
    if kupiec_95 is None:
        kupiec_95 = metrics.kupiec_backtest(ptf, 0.95)
    if kupiec_99 is None:
        kupiec_99 = metrics.kupiec_backtest(ptf, 0.99)
    if ms_ptf is None:
        ms_ptf = metrics.fit_markov_switching(ptf)

    _build_synthese(wb, ptf, bench, risk_free_rate, garch_ptf, kupiec_95, kupiec_99)
    _build_series(wb, portfolio, ptf, bench)
    _build_garch(wb, ptf, bench, garch_ptf, garch_bench)
    _build_regimes(wb, ptf, ms_ptf)
    _build_rolling_sharpe(wb, ptf, bench, risk_free_rate)
    _build_correlations(wb, portfolio)
    _build_contribution(wb, portfolio)
    _build_distribution(wb, ptf, bench)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
